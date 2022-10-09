'''
 daniel testa . daniel.testa.t@gmail.com
 ---- versión 1.33
 6 de octubre de 2022

 Script para generar archivos xlsx nuevos a partir de dos
 números de remito que definen una pila

 PARA HACER:
 - Administrar mejor los errores try/exept
 - Crear lógica flexible para cuando son más de 2 remitos
'''
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.worksheet.page import PageMargins, PrintPageSetup
from operator import index
import PySimpleGUI as sg
import os


def generar_pila(arch_ent, fecha_reunion, pila_destino, remito_1, remito_2):

    a_e = arch_ent
    f_r = fecha_reunion
    p_r = pila_destino
    r1 = int(remito_1)
    r2 = int(remito_2)

    libro_ent = openpyxl.load_workbook(a_e)
    hoja_ent = libro_ent['Hoja1']
    celda_ent = hoja_ent.cell
    #------------------ Crear nuevo libro de excel
    ruta = os.getcwd()
    arch_sal = f'{ruta}\\REUNION {f_r}\\{p_r} __{f_r}.xlsx'
    libro_nuevo = openpyxl.Workbook()
    libro_nuevo.save(arch_sal)
    libro_sal = openpyxl.load_workbook(arch_sal)
    hoja_sal = libro_sal['Sheet']
    hoja_sal.title ='sobres y acuses a imprimir'
    celda_sal = hoja_sal.cell
    #------------------ Opciones de impresión ----------------(paperSize= 9 == A4. Modulo worksheet.py línea 88)
    hoja_sal.page_setup = PrintPageSetup(orientation = 'landscape', paperSize= 9, scale= 84)
    hoja_sal.page_margins = PageMargins(left=0, right=0, top=0, bottom=0, header=0, footer=0)
    #------------------  Definición de estilos
    letra_encabezaado = Font(name='Candara', size=10, color='00FFFFFF')
    letra_común = Font(name='Arial', size=10, color='00000000')
    fondoNegro = PatternFill(start_color='00000000', end_color='00000000', fill_type='solid')
    borde_abajo = Border(bottom=Side(border_style='thin' ,color='FF000000'))
    #------------------ Encabezados
    hoja_sal['A1'] = 'CU - C.A.'
    hoja_sal['B1'] = 'BENEFICIARIO'
    hoja_sal['C1'] = 'CAUSANTE'
    hoja_sal['D1'] = 'DOMICILIO'
    hoja_sal['E1'] = 'DOMICILIO2'
    hoja_sal['F1'] = 'LOCALIDAD'
    hoja_sal['G1'] = 'C. P.'
    hoja_sal['H1'] = 'PROVINCIA'
    hoja_sal['I1'] = 'EXPEDIENTE'
    hoja_sal['J1'] = 'FECHA RES'
    #------------------ Anchos de columnas
    hoja_sal.column_dimensions['A'].width = 25
    hoja_sal.column_dimensions['B'].width = 27
    hoja_sal.column_dimensions['C'].width = 18
    hoja_sal.column_dimensions['D'].width = 27
    hoja_sal.column_dimensions['E'].width = 11
    hoja_sal.column_dimensions['F'].width = 20
    hoja_sal.column_dimensions['G'].width = 5
    hoja_sal.column_dimensions['H'].width = 5
    hoja_sal.column_dimensions['I'].width = 15
    hoja_sal.column_dimensions['H'].width = 12
    hoja_sal.column_dimensions['K'].width = 8

    #------------------ Estilo para encabezado
    for columna in range(1, 11):
        celda_sal(row=1, column=columna).fill = fondoNegro
        celda_sal(row=1, column=columna).font = letra_encabezaado

    #------------------ Encontrar número de fila en que aparece por primera vez el remito_1
    for _ in range(2,hoja_ent.max_row+1):
        if celda_ent(row=_, column=11).value == r1:
            primera_fila_ent_1 = index(_)
            primera_fila_1 = primera_fila_ent_1 - 2
            break

    #------------------ Encontrar número de fila en que aparece por ultima vez el remito_1
    lista_a_medir = []
    for _ in range(2,hoja_ent.max_row+1):
        if celda_ent(row=_, column=11).value == r1:
            lista_a_medir.append(index(_))
    lista_a_medir.reverse()
    ultima_fila_1 = lista_a_medir[0]

    #------------------ Encontrar número de fila en que aparece por primera vez el remito_2
    for _ in range(2,hoja_ent.max_row+1):
        if celda_ent(row=_, column=11).value == r2:
            primera_fila_ent_2 = index(_)
            primera_fila_2 = primera_fila_ent_2 - 1
            break

    #------------------ Copiar en la pila el primer remito
    for i in range(2, hoja_ent.max_row+1):
        if celda_ent(row=i, column=11).value == r1:
            for j in range(2, 12):
                celda_sal(row=i-primera_fila_1, column=j).value = celda_ent(row=i, column=j).value

    #------------------ Copiar en la pila el segundo remito
    dif_ent_1_y_2 = ultima_fila_1 - primera_fila_1
    for i in range(2, hoja_ent.max_row+1):
         if celda_ent(row=i, column=11).value == r2:
             for j in range(2, 12):
                 celda_sal(row=i - primera_fila_2 + dif_ent_1_y_2, column=j).value = celda_ent(row=i, column=j).value

    #------------------ Estilo común. Al final para tener hoja_sal.max_row .max_column
    for columna in range(1, hoja_sal.max_column):
        for fila in range(2, hoja_sal.max_row+1):
            celda_sal(row=fila, column=columna).font = letra_común
            celda_sal(row=fila, column=columna).border = borde_abajo
            hoja_sal.row_dimensions[fila].height = 20

    #------------------ Cambiar 'Fecha' en columna 'FECHA RES' por 'Fecha pila'
    for fila in range(2, hoja_sal.max_row+1):
        celda_sal(row=fila, column=10).value = f'{f_r} {p_r}'

    libro_sal.save(arch_sal)

def const_lista_remitos(arch_ent):
    a_e = arch_ent
    libro_ent = openpyxl.load_workbook(a_e)
    hoja_ent = libro_ent['Hoja1']
    celda_ = hoja_ent.cell
    n_remitos = []

    for _ in range(2, hoja_ent.max_row+1):
        n_remitos.append( celda_(row=_,column=11).value)
        unicos = list(set(n_remitos))
    return unicos


def main():
    sg.theme('SystemDefaultForReal')
    diseño = [ [sg.Text('Elegir archivo Excel listado completo de reunión')],
                [sg.Text('Elegir archivo:'), sg.Input(key='xlsx_entrada' ,change_submits=True), sg.FileBrowse(key='xlsx_entrada', button_text= 'Buscar', file_types=(('Archivos Excel', '*.xlsx'),))],
                [sg.Text('Ingresar fecha de reunión (dd-mm-aa)')],
                [sg.Text('Fecha:          '), sg.Input(key='fecha_r', change_submits=True)],
                [sg.Text('Ingresar pila (p1, p2, p..)')],
                [sg.Text('Pila:              '), sg.Input(key='pila_', change_submits=True)],
                [sg.Text('Ingresar número del primer remito de la pila')],
                [sg.Text('Remito 1:      '), sg.Input(key='remito_1', change_submits=True)],
                [sg.Text('Ingresar número del segundo remito de la pila')],
                [sg.Text('Remito 2:      '), sg.Input(key='remito_2', change_submits=True)],
                [sg.Button('Aceptar', expand_x=True), sg.Button('Otra Pila', expand_x=True), sg.Button('Salir', expand_x=True)]   ]

    window = sg.Window('Completar los datos necesarios para generar las pilas', diseño)

    while True:
        event, values = window.read()

        arch_ent = values['xlsx_entrada']
        fecha_reunion = values['fecha_r']
        pila_destino = values['pila_']
        remito_1 = values['remito_1']
        remito_2 = values['remito_2']
        ruta_de_archivo = os.path.realpath(arch_ent)
        carpeta_de_trabajo = os.path.dirname(ruta_de_archivo)
        os.chdir(carpeta_de_trabajo)

        if event == sg.WIN_CLOSED or event == 'Salir':
            break
        elif event == 'Otra Pila':
            window['pila_']('')
            window['remito_1']('')
            window['remito_2']('')
        elif event == 'Aceptar':
            num_r = const_lista_remitos(arch_ent)
            rem1_str = remito_1
            rem2_str = remito_2
            ''' print(num_r) # Utilizado para chequear el contenido de la lista devuelta por “cons_lista_remitos()”'''

            if arch_ent != '' and fecha_reunion != '' and pila_destino != '' and remito_1 != '' and remito_2 != '':

                if rem1_str.isnumeric() and rem2_str.isnumeric():

                    if int(remito_1) in num_r:
                        dir_reunión = 'REUNION ' + fecha_reunion
                        existe_dir_reunión = os.path.isdir(dir_reunión)

                        if not existe_dir_reunión:
                            elección = sg.popup_yes_no(f'La carpeta {dir_reunión} no existe.\n¿Querés crearla?')
                            if elección == 'No':
                                break
                            elif elección == 'Yes':
                                os.makedirs(dir_reunión)
                        else:
                            sg.popup_auto_close(f'Las pilas se guardarán en la carpeta \'{dir_reunión}\'.')

                        generar_pila(arch_ent, fecha_reunion, pila_destino, remito_1, remito_2)
                        window['pila_']('')
                        window['remito_1']('')
                        window['remito_2']('')

                        if os.path.isfile(f'{dir_reunión}\\{pila_destino} __{fecha_reunion}.xlsx'):
                            sg.popup_auto_close(f'El archivo \'{pila_destino} __{fecha_reunion}.xlsx\' fue creado.')
                    else:
                        sg.popup_auto_close(f'El número de remito \'{remito_1}\' no existe en esta REUNIÓN:')
                        window['remito_1']('')
                else:
                    sg.popup(f'El número de remito ingresado, no es un número.')
                    window['remito_1']('')
                    window['remito_2']('')
            else:
                sg.popup('Completar todos los campos.')
                window['remito_1']('')
                window['remito_2']('')
    window.Close()


if __name__ == '__main__':
    main()
