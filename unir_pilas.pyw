"""
 daniel testa . daniel.testa.t@gmail.com
 ---- versión 1.0
 11 de octubre de 2022

Script para concatenar pilas (archivos excel). La primera fila se toma
solo de la primer pila para tener los encabezados.
Este script existe exclusivamente a lo fines de generar un único listado e
imprimir los ARs y Sobres de todas las pilas o al menos varias, de una vez.
"""

import openpyxl
import os
import PySimpleGUI as sg


def main():
    while True:
        archivos = popup_buscar_archivos(
                   'Seleccionar pilas a combinar para imprimir ARs y '
                   + 'Sobres e ingresar archivo de sallida.',
                   title='Buscar pilas e ingresar archivo de salida')
        if not archivos:
            break
        else:
            lista_de_xlsx = archivos[0].split(";")
            concatenar_xlsx(lista_de_xlsx).save(archivos[1])
            return False


def popup_buscar_archivos(mensage, title=None):
    sg.theme("SystemDefaultForReal")
    diseño = [[sg.Text(mensage)],
              [sg.Text('Pilas xlsx          '), sg.Input(key='-PILAS-'), sg.FilesBrowse('Buscar')],
              [sg.Text('Archivo de salida'), sg.Input(default_text='salida.xlsx', key='-ARCH_SAL-')],
              [sg.Button('Aceptar'), sg.Button('Cancelar')]]
    window = sg.Window(title if title else mensage, diseño)
    while True:
        event, values = window.read(close=True)
        if event == sg.WIN_CLOSED or event == 'Cancelar':
            break
        elif (event == 'Aceptar' and values['-PILAS-'] != 'None' and
                values['-ARCH_SAL-']):
            return [values['-PILAS-'], values['-ARCH_SAL-']]
        else:
            sg.popup('Completar todos los campos')
    window.close()


def concatenar_xlsx(lista_de_xlsx):
    pilas = []
    for xlsx in lista_de_xlsx:
        libro = openpyxl.load_workbook(xlsx)
        pilas.append(libro)
    libro_sal = openpyxl.Workbook()
    hoja_sal = libro_sal.worksheets[0]
    #--------- copiar  Encabezados
    libro1 = pilas[0]
    hoja1 = libro1.worksheets[0]
    for j in range(1, hoja1.max_column +1):
        hoja_sal.cell(row=1, column=j).value = hoja1.cell(row=1, column=j).value
    #--------- Concatenar todas las pilas dejando fuera los encabezados.
    fila_actual = 2
    for libro in pilas:
        for hojas in libro.worksheets:
            max_f =  hojas.max_row
            max_c = hojas.max_column
            for i in range(2, max_f + 1):
                for j in range(1, max_c + 1):
                    celda_actual = hojas.cell(row=i, column=j)
                    hoja_sal.cell(row=fila_actual, column=j).value = celda_actual.value
                fila_actual += 1
    return libro_sal


if __name__ == '__main__':
    main()
