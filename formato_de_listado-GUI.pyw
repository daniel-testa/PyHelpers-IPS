'''
daniel testa. daniel.testa.t@gmail.com
6 de octubre de 2022
---- versión 1.33

Script para automatizar los arreglos, modificaciones, etc en el archivo
xlsx generado al importar los datos proveeidos por computación en el archivo
'_FREE4_WKF_EXPEDIENTES_REUNION_NOTIFICACIONES_X_MAIL.AAAA-MM-DD.TXT'
a la plantilla 'Plantilla_COMPLETO.xlsx'
'''

import PySimpleGUI as sg
import openpyxl
from openpyxl.styles import numbers, PatternFill

def ciudad (celda_):
    conteo_fila = 1
    while True:
        conteo_fila +=1
        col_ciudad = celda_(row=conteo_fila, column=6).value
        if col_ciudad != None:
            celda_(row=conteo_fila, column=6).value =(
                    col_ciudad.replace('CAPITAL FEDERAL', 'CABA')
                    .replace('ALMIRANTE BROWN', 'ADROGUÉ')
                    .replace('BUENOS AIRES', 'BS. AS.')
                    .replace('BELEN DE ESCOBAR', 'B DE ESCOBAR')
                    .replace('CARMEN DE PATAGONES', 'PATAGONES')
                    .replace('CORONEL BRANDSEN', 'BRANDSEN')
                    .replace('CIUDAD JARDIN DEL PALOMAR', 'C J DEL PALOMAR')
                    .replace('FLORENTINO AMEGHINO', 'F AMEGHINO')
                    .replace('GENERAL LAS HERAS', 'GRAL LAS HERAS')
                    .replace('GENERAL LAMADRID', 'GRAL LAMADRID')
                    .replace('GENERAL SAN MARTIN', 'GRAL SAN MARTIN')
                    .replace('GREGORIO DE LAFERRERE', 'LAFERRERE')
                    .replace('JOSE CLEMENTE PAZ', 'JOSE C PAZ')
                    .replace('LISANDRO OLMOS ETCHEVERRY', 'LISANDRO OLMOS')
                    .replace('NORBERTO DE LA RIESTRA', 'N DE LA RIESTRA')
                    .replace('REMEDIOS DE ESCALADA', 'R DE ESCALADA')
                    .replace('SAN ANDRES DE GILES', 'GILES')
                    .replace('SAN ANTONIO DE PADUA', 'SAN A DE PADUA')
                    .replace('SAN ANTONIO DE ARECO', 'SAN A DE ARECO')
                    .replace('SAN MIGUEL DEL MONTE', 'MONTE')
                    .replace('SAN NICOLAS DE LOS ARROYOS', 'SAN NICOLAS')
                    .replace('SIERRA DE LA VENTANA', 'S DE LA VENTANA')
                    .replace('VILLA SANTOS TESEI', 'VILLA TESEI')
                    )
        else:
            break

def provincia (celda_):
    conteo_fila = 1
    while True:
        conteo_fila +=1
        col_pcia = celda_(row=conteo_fila, column=8).value
        if col_pcia != None:
            celda_(row=conteo_fila, column=8).value = (
                    col_pcia.replace('CAPITAL FEDERAL', 'CABA')
                    .replace('BUENOS AIRES', 'BS. AS.')
                    )
        else:
           break

def expediente (celda_):
    conteo_fila = 1
    while True:
        conteo_fila +=1
        col_exte = celda_(row=conteo_fila, column=9).value
        if col_exte != None:
            celda_(row=conteo_fila, column=9).value = (
                    col_exte.strip('-000')
                    .replace('-0-', '-')
                    .replace('-00', '-')
                    .replace('-0', '-')
                    )
        else:
           break

def fecha (celda_, fecha_reunion):
    conteo_fila = 1
    while True:
        conteo_fila +=1
        col_fecha = celda_(row=conteo_fila, column=10).value
        if col_fecha != None:
            celda_(row=conteo_fila, column=10).number_format = numbers.FORMAT_TEXT
            celda_(row=conteo_fila, column=10).value = fecha_reunion
        else:
           break

def extranjera (celda_, planilla_de_trabajo):
    conteo_fila = 1
    while True:
        conteo_fila +=1
        fondoRojo = PatternFill(
                start_color='FFFF0000',
                end_color='FFFF0000', fill_type='solid'
                )
        col_extranjero = celda_(row=conteo_fila, column=8).value
        if col_extranjero != None:
            if (celda_(row=conteo_fila, column=8).value.__contains__('OTRA / NC') or
                celda_(row=conteo_fila, column=8).value.isspace()):
                for i in range(1, planilla_de_trabajo.max_column):
                    celda_(row=conteo_fila, column=i).fill = fondoRojo
        else:
           break

def lista_imprimir (celda_, listado_de_remitos, planilla_de_trabajo, f):

    listado_de_remitos['A1'] = f'Reunión {f} - {str(planilla_de_trabajo.max_row - 2)} expedientes.'
    n_remitos = []
    for _ in range(2, planilla_de_trabajo.max_row):
        n_remitos.append( celda_(row=_,column=11).value)
        unicos = sorted(list(set(n_remitos)))

    for remito in unicos:
       listado_de_remitos.append([remito])


def main():
    sg.theme('SystemDefaultForReal')
    diseño = [[sg.Text('Elegir archivo Excel a modificar')],
              [sg.Text('Elegir archivo:'), sg.Input(key='xlsx_entrada' ,change_submits=True), sg.FileBrowse(key='xlsx_entrada', button_text= 'Buscar', file_types=(('Archivos Excel', '*.xlsx'),))],
              [sg.Text('Ingresar fecha de reunión (dd-mm-aa)')],
              [sg.Text('Fecha:           '), sg.Input(key='fecha_r', change_submits=True)],
              [sg.Button('Aceptar', expand_x=True), sg.Button('Salir', expand_x=True), sg.Button('Limpiar', expand_x=True)]]
    window = sg.Window('Buscar archivo e ingresar fecha', diseño)

    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == 'Salir':
            break
        elif event == 'Limpiar':
            window['xlsx_entrada']('')
            window['fecha_r']('')
        elif event == 'Aceptar':

            nombre_archivo_ent = values['xlsx_entrada']
            fecha_reunion = values['fecha_r']
            nombre_archivo_sal = nombre_archivo_ent
            datos = openpyxl.load_workbook(nombre_archivo_ent)
            planilla_de_trabajo = datos['Hoja1']
            celda_ = planilla_de_trabajo.cell
            listado_de_remitos = datos['Remitos y Fechas']

            ciudad(celda_)
            provincia(celda_)
            expediente(celda_)
            fecha(celda_, fecha_reunion)
            extranjera(celda_, planilla_de_trabajo)
            lista_imprimir(celda_, listado_de_remitos, planilla_de_trabajo, fecha_reunion)
            datos.save(nombre_archivo_sal)
            sg.PopupOK('Exito')
    window.close()

if __name__ == '__main__':
    main()
